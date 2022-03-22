import openpyxl
import os
excel_local ='../new.xlsx'
def excel(value,time,data):
    data = int(data)
    value = (value-1)*2+1
    file_path = os.path.dirname(os.path.abspath(__file__))
    base_path = os.path.join(file_path, excel_local)
    workbook = openpyxl.load_workbook(base_path)
    worksheet = workbook.get_sheet_by_name('工作表1')
    i = 1
    while(worksheet.cell(row=i,column=value).value!=None):
        i+=1
    print(i)
    workbook = openpyxl.load_workbook(excel_local)
    sheet = workbook.worksheets[0]
    it = str(chr(value+64)+str(i))
    #print(it)
    sheet[it]=time
    it = str(chr(value+65)+str(i))
    #print(it)
    sheet[it]=data
    workbook.save(excel_local)
    print("run over")
    print("value",value)
    print("time",time)
    print("data",data)
    print("------------------------")